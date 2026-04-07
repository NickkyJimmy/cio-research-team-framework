#!/usr/bin/env python3
"""Build IRIS SURVEY DETAIL sheet from a simple JSON spec.

Usage:
  python build_iris_survey_detail.py --template /path/IRIS_Survey_Template.xlsx \
    --spec /path/questions.json --output /path/output.xlsx

Spec format (JSON):
{
  "sheet_name": "SURVEY DETAIL",
  "pages": [
    {
      "name": "PAGENAME: SINGLE_ANSWER & MULTIPLE_ANSWER",
      "questions": [
        {
          "id": "Q1",
          "text": "Câu hỏi ...",
          "type": "SINGLE_ANSWER",
          "instruction": "Vui lòng chọn 1 phương án",
          "options": ["A", "B", "C"],
          "special_codes": {"Khác, ghi rõ": "other", "Tôi không biết": "none"}
        }
      ]
    }
  ]
}
"""

import argparse
import json
import shutil
from openpyxl import load_workbook


def _append_row(ws, row_idx, values):
    for col_idx, val in enumerate(values, start=1):
        ws.cell(row=row_idx, column=col_idx, value=val)
    return row_idx + 1


def _find_last_row(ws):
    last = ws.max_row
    while last > 1 and all(ws.cell(row=last, column=c).value is None for c in range(1, 10)):
        last -= 1
    return last


def _find_first_pagename_row(ws):
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and val.strip().startswith("PAGENAME"):
            return r
    return None


def _add_question_block(ws, row_idx, q):
    qid = q["id"]
    text = q["text"]
    qtype = q["type"]
    instruction = q.get("instruction", "")
    options = q.get("options", [])
    special_codes = q.get("special_codes", {})
    data_type = q.get("data_type", "Text")

    # Question header
    row_idx = _append_row(ws, row_idx, [qid, "Nội dung câu hỏi", text, ""])
    row_idx = _append_row(ws, row_idx, ["", "Loại câu", qtype, ""])

    if qtype == "OPEN_ENDED":
        row_idx = _append_row(ws, row_idx, ["", "Dạng", data_type, ""])
        row_idx = _append_row(ws, row_idx, ["", "Hướng dẫn trả lời", "Điền câu trả lời vào ô trống", ""])
        return row_idx

    row_idx = _append_row(ws, row_idx, ["", "Hướng dẫn trả lời", instruction, ""])
    row_idx = _append_row(ws, row_idx, ["", "Code", "Option", ""])

    code = 1
    for opt in options:
        row = ["", code, opt, ""]
        # Special option codes go in column D (index 4)
        if opt in special_codes:
            row[3] = special_codes[opt]
        row_idx = _append_row(ws, row_idx, row)
        code += 1

    return row_idx


def build(template_path, spec_path, output_path):
    with open(spec_path, "r", encoding="utf-8") as f:
        spec = json.load(f)

    shutil.copyfile(template_path, output_path)
    wb = load_workbook(output_path)

    sheet_name = "SURVEY DETAIL"
    if sheet_name not in wb.sheetnames:
        raise ValueError("Template file is missing required sheet: SURVEY DETAIL")
    ws = wb[sheet_name]

    # Optionally keep only SURVEY DETAIL
    if spec.get("keep_only_survey_detail"):
        for name in list(wb.sheetnames):
            if name != sheet_name:
                del wb[name]

    # Optional survey objective block
    survey_objective = spec.get("survey_objective")
    if survey_objective:
        ws.cell(row=1, column=1, value="SURVEY OBJECTIVE")
        ws.cell(row=2, column=2, value="Context / Background")
        ws.cell(row=2, column=3, value=survey_objective.get("context"))
        ws.cell(row=3, column=2, value="Survey Objectives")
        ws.cell(row=3, column=3, value=survey_objective.get("objectives"))
        ws.cell(row=4, column=2, value="Target Audience")
        ws.cell(row=4, column=3, value=survey_objective.get("audience"))

    pagename_row = _find_first_pagename_row(ws)
    if pagename_row is not None:
        row_idx = pagename_row
    else:
        row_idx = _find_last_row(ws) + 2

    for page in spec.get("pages", []):
        row_idx = _append_row(ws, row_idx, [page["name"]])
        for q in page.get("questions", []):
            row_idx = _add_question_block(ws, row_idx, q)

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Build IRIS SURVEY DETAIL sheet from JSON spec.")
    parser.add_argument("--template", required=True, help="Path to IRIS_Survey_Template.xlsx")
    parser.add_argument("--spec", required=True, help="Path to questions JSON spec")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    build(args.template, args.spec, args.output)


if __name__ == "__main__":
    main()
